#信息验证
# 定义主函数
import datetime

_EXPIRY_DATE = datetime.date(2026, 5, 1)


def main():
    if datetime.date.today() >= _EXPIRY_DATE:
        return

# 确保当此脚本作为主程序运行时，调用main函数
if __name__ == "__main__":
    main()


# 主代码
import tkinter as tk
from tkinter import messagebox
from decimal import Decimal, ROUND_HALF_UP
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
import re
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import math
import os
import subprocess
import sys


def get_fill_signature(cell):
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return None
    fg = getattr(fill.fgColor, "value", None)
    bg = getattr(fill.bgColor, "value", None)
    return (fill.patternType, fg, bg)


# 横版分数汇总
def summarize_data_horizontal():
    # 打开分数excel工作簿
    workbook = openpyxl.load_workbook('工作簿1.xlsx')
    sheet = workbook[workbook.sheetnames[0]]

    # 定义变量保存两个不同颜色单元格的坐标
    cell1_coord = None
    cell2_coord = None

    # 遍历表格，找到颜色不同的两个单元格
    cell1_signature = None
    for row in sheet.iter_rows():
        for cell in row:
            signature = get_fill_signature(cell)
            if signature is None:
                continue
            if cell1_coord is None:
                cell1_coord = cell.coordinate
                cell1_signature = signature
            elif signature != cell1_signature:
                cell2_coord = cell.coordinate
                break
        if cell2_coord is not None:
            break

    # 输出两个不同颜色单元格的坐标
    print("颜色单元格1坐标:", cell1_coord)
    print("颜色单元格2坐标:", cell2_coord)
    if cell1_coord is None or cell2_coord is None:
        messagebox.showinfo("提示", "未找到两种不同颜色的单元格，无法继续汇总。")
        return

    # 获取评审项列表及其坐标
    review_items = []
    review_items_coords = []
    cell1_column_letter, cell1_row_number = coordinate_from_string(cell1_coord)
    start_column_index = openpyxl.utils.column_index_from_string(cell1_column_letter)
    end_column_index = start_column_index
    row_index = cell1_row_number
    while sheet.cell(row=row_index, column=end_column_index).value is not None:
        cell = sheet.cell(row=row_index, column=end_column_index)
        review_items.append(cell.value)
        review_items_coords.append(cell.coordinate)
        end_column_index += 1
    print("评审项:", review_items)
    print("评审项坐标:", review_items_coords)

    # 获取公司名称列表及其坐标
    company_names = []
    company_names_coords = []
    cell2_column_letter, cell2_row_number = coordinate_from_string(cell2_coord)
    column_index = openpyxl.utils.column_index_from_string(cell2_column_letter)
    row_index = cell2_row_number
    while sheet.cell(row=row_index, column=column_index).value is not None:
        cell = sheet.cell(row=row_index, column=column_index)
        company_names.append(cell.value)
        company_names_coords.append(cell.coordinate)
        row_index += 1

    # 输出公司名称数量和评审项目数量
    print("评审项目数量:", len(review_items))
    print("公司名称数量:", len(company_names))

    # 创建汇总分析表
    summary_workbook = openpyxl.Workbook()
    summary_sheet = summary_workbook.active

    # 调整列宽和行高
    summary_sheet.column_dimensions['A'].width = 20
    summary_sheet.column_dimensions['C'].width = 15
    summary_sheet.column_dimensions['D'].width = 100
    for row in range(1, 5001):
        summary_sheet.row_dimensions[row].height = 200

    # 写入列抬头
    summary_sheet['A1'] = '公司名称'
    summary_sheet['B1'] = '评审内容'
    summary_sheet['C1'] = '评审因素'
    summary_sheet['D1'] = '评审细则'
    summary_sheet['E1'] = '分值'
    summary_sheet['F1'] = '主客观分'

    # 写入每个表单的列抬头
    for i, sheet_name in enumerate(workbook.sheetnames):
        column_index = i + 7  # 从F列开始，列索引依次增加
        column_letter = openpyxl.utils.get_column_letter(column_index)
        summary_sheet[column_letter + '1'] = sheet_name

    # 写入公司名称和评审项
    summary_row = 2
    for company_name, company_coord in zip(company_names, company_names_coords):
        for review_item, review_coord in zip(review_items, review_items_coords):
            summary_sheet.cell(row=summary_row, column=1).value = company_name
            summary_sheet.cell(row=summary_row, column=3).value = review_item

            # 获取评审内容、评审细则和评审项分值的单元格坐标
            review_coord_column_letter = ''.join(filter(str.isalpha, review_coord))  # 提取列字母
            cell2_coord_row_number = int(''.join(filter(str.isdigit, cell2_coord)))  # 提取行数字
            value_cell_coord0 = f'{review_coord_column_letter}{cell2_coord_row_number - 4}'  # 评审内容
            value_cell_coord1 = f'{review_coord_column_letter}{cell2_coord_row_number - 2}'  # 评审细则
            value_cell_coord = f'{review_coord_column_letter}{cell2_coord_row_number - 1}'  # 评审项分值

            # 获取评审内容、评审细则和评审项分值
            value0 = sheet[value_cell_coord0].value
            value1 = sheet[value_cell_coord1].value
            value = sheet[value_cell_coord].value

            # 在B、D、E列填入评审细则和评审项分值
            summary_sheet.cell(row=summary_row, column=2).value = value0
            summary_sheet.cell(row=summary_row, column=4).value = value1
            summary_sheet.cell(row=summary_row, column=5).value = value

            # 写入每个表单的评审项分数
            company_coord_row_number = int(''.join(filter(str.isdigit, company_coord)))  # 提取行数字
            for i, sheet_name in enumerate(workbook.sheetnames):
                column_index = i + 7
                column_letter = openpyxl.utils.get_column_letter(column_index)
                score_cell_coord = f'{review_coord_column_letter}{company_coord_row_number}'

                # 获取评审项分数
                score = workbook[sheet_name][score_cell_coord].value

                # 在对应的列填入评审项分数
                summary_sheet.cell(row=summary_row, column=column_index).value = score

            summary_row += 1

    # 调整列宽和行高
    summary_sheet.column_dimensions['D'].width = 30
    for row in range(1, 5001):
        summary_sheet.row_dimensions[row].height = 13.5

    # 保存汇总分析表
    summary_workbook.save('汇总分析表.xlsx')
    print("汇总分析表已生成。")
    messagebox.showinfo("提示", "横版数据汇总完成！")


# 纵版分数汇总
def summarize_data_vertical():
    # 打开分数excel工作簿
    workbook = openpyxl.load_workbook('工作簿1.xlsx')
    sheet = workbook[workbook.sheetnames[0]]

    # 定义变量保存两个不同颜色单元格的坐标
    color_cell_coord_1 = None
    color_cell_coord_2 = None

    # 遍历表格，找到颜色不同的两个单元格
    color1_signature = None
    for row in sheet.iter_rows():
        for cell in row:
            signature = get_fill_signature(cell)
            if signature is None:
                continue
            if color_cell_coord_1 is None:
                color_cell_coord_1 = cell.coordinate
                color1_signature = signature
            elif signature != color1_signature:
                color_cell_coord_2 = cell.coordinate
                break
        if color_cell_coord_2 is not None:
            break

    # 输出两个不同颜色单元格的坐标
    print("颜色单元格1坐标:", color_cell_coord_1)
    print("颜色单元格2坐标:", color_cell_coord_2)
    if color_cell_coord_1 is None or color_cell_coord_2 is None:
        messagebox.showinfo("提示", "未找到两种不同颜色的单元格，无法继续汇总。")
        return

    # 获取公司名称列表及其坐标
    color1_column_letter, color1_row_number = coordinate_from_string(color_cell_coord_1)
    start_column_index = openpyxl.utils.column_index_from_string(color1_column_letter)
    revcompany_names = [cell[0].value for cell in sheet.iter_cols(min_row=color1_row_number, max_row=color1_row_number, min_col=start_column_index) if cell[0].value is not None]
    revcompany_names_coords = [cell[0].coordinate for cell in sheet.iter_cols(min_row=color1_row_number, max_row=color1_row_number, min_col=start_column_index) if cell[0].value is not None]
    print("公司名称数量:", len(revcompany_names))
    print("公司名称列表", revcompany_names)
    print("公司名称坐标", revcompany_names_coords)

    # 获取评审项列表及其坐标
    color2_column_letter, color2_row_number = coordinate_from_string(color_cell_coord_2)
    column_index = openpyxl.utils.column_index_from_string(color2_column_letter)
    review_items = [cell[0].value for cell in sheet.iter_rows(min_row=color2_row_number, min_col=column_index, max_col=column_index) if cell[0].value is not None]
    review_items_coords = [cell[0].coordinate for cell in sheet.iter_rows(min_row=color2_row_number, min_col=column_index, max_col=column_index) if cell[0].value is not None]
    print("评审项目数量:", len(review_items))
    print("评审项", review_items)
    print("评审项坐标", review_items_coords)

    # 创建汇总分析表
    summary_workbook = openpyxl.Workbook()
    summary_sheet = summary_workbook.active

    # 调整列宽和行高
    summary_sheet.column_dimensions['A'].width = 20
    summary_sheet.column_dimensions['C'].width = 15
    summary_sheet.column_dimensions['D'].width = 100
    for row in range(1, 5001):
        summary_sheet.row_dimensions[row].height = 200

    # 写入列抬头
    summary_sheet['A1'] = '公司名称'
    summary_sheet['B1'] = '评审内容'
    summary_sheet['C1'] = '评审因素'
    summary_sheet['D1'] = '评审细则'
    summary_sheet['E1'] = '分值'
    summary_sheet['F1'] = '主客观分'

    # 写入每个表单的列抬头
    for i, sheet_name in enumerate(workbook.sheetnames):
        column_letter = openpyxl.utils.get_column_letter(i + 7)  # 从F列开始，列索引依次增加
        summary_sheet[column_letter + '1'] = sheet_name

    # 写入评审项和公司名称
    summary_row = 2
    for review_item, review_coord in zip(revcompany_names, revcompany_names_coords):
        for company_name, company_coord in zip(review_items, review_items_coords):
            summary_sheet.cell(row=summary_row, column=1).value = review_item
            summary_sheet.cell(row=summary_row, column=3).value = company_name

            # 获取评审内容、评审细则和评审项分值的单元格坐标
            color_cell_coord_1_column_letter = ''.join(filter(str.isalpha, color_cell_coord_1))  # 提取颜色1的列字母
            company_coord_row_number = int(''.join(filter(str.isdigit, company_coord)))  # 提取公司名称行数字
            value_cell_coord0 = f'{openpyxl.utils.get_column_letter(openpyxl.utils.column_index_from_string(color_cell_coord_1_column_letter) - 4)}{company_coord_row_number}'
            value_cell_coord1 = f'{openpyxl.utils.get_column_letter(openpyxl.utils.column_index_from_string(color_cell_coord_1_column_letter) - 2)}{company_coord_row_number}'
            value_cell_coord = f'{openpyxl.utils.get_column_letter(openpyxl.utils.column_index_from_string(color_cell_coord_1_column_letter) - 1)}{company_coord_row_number}'

            # 获取评审内容、评审细则、评审项分值
            value0 = sheet[value_cell_coord0].value
            value1 = sheet[value_cell_coord1].value
            value = sheet[value_cell_coord].value

            # 在B、D、E列填入评审项分值
            summary_sheet.cell(row=summary_row, column=2).value = value0
            summary_sheet.cell(row=summary_row, column=4).value = value1
            summary_sheet.cell(row=summary_row, column=5).value = value

            # 写入每个表单的评审项分数
            review_coord_column_letter = ''.join(filter(str.isalpha, review_coord))  # 提取列字母
            for i, sheet_name in enumerate(workbook.sheetnames):
                column_index = i + 7
                column_letter = openpyxl.utils.get_column_letter(column_index)
                score_cell_coord = f'{review_coord_column_letter}{company_coord_row_number}'

                # 获取评审项分数
                score = workbook[sheet_name][score_cell_coord].value

                # 在对应的列填入评审项分数
                summary_sheet.cell(row=summary_row, column=column_index).value = score

            summary_row += 1

    # 调整列宽和行高
    summary_sheet.column_dimensions['D'].width = 30
    for row in range(1, 5001):
        summary_sheet.row_dimensions[row].height = 13.5

    # 保存汇总分析表
    summary_workbook.save('汇总分析表.xlsx')
    print("汇总分析表已生成。")
    messagebox.showinfo("提示", "纵版数据汇总完成！")


#主客观分析
def subjective_analysis():

        # 重新加载保存的汇总分析表文件
    summary_workbook = openpyxl.load_workbook('汇总分析表.xlsx')
    summary_sheet = summary_workbook.active

    # 删除D列单元格值包含的中英文空格
    for row in summary_sheet.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            if cell.value:
                cell.value = cell.value.replace(" ", "")

    import re

    # 判断D列单元格内容，在F列填入"客观分"或"主观分"
    row_index = 2
    while summary_sheet.cell(row=row_index, column=4).value:
        cell_c = summary_sheet.cell(row=row_index, column=4)
        cell_e = summary_sheet.cell(row=row_index, column=6)

        pattern = r'(加|得|扣|减|少|加0.|得0.|扣0.|减0.|少0.|加1.|得1.|扣1.|减1.|少1.|加2.|得2.|扣2.|减2.|少2.|加3.|得3.|扣3.|减3.|少3.)(\d+)分'
        match = re.search(pattern, cell_c.value)  
        if match:
            cell_e.value = "客观分"
        else:
            cell_e.value = "主观分"

        row_index += 1


    # 获取C列的最大行号
    c_max_row = summary_sheet.max_row

    # 从B2单元格开始向下
    for i in range(2, c_max_row):
        current_cell = summary_sheet.cell(row=i, column=2)  # 获取当前单元格
        if current_cell.value is None:  # 如果单元格为空
            previous_cell = summary_sheet.cell(row=i-1, column=2)  # 获取上一个非空单元格
            current_cell.value = previous_cell.value  # 填充上一个非空单元格的内容

    # 保存汇总分析表
    summary_workbook.save('汇总分析表.xlsx')
    print("主客观分析完成。")
    messagebox.showinfo("提示", "主客观分判断完成！")


# 开始分值校验
def analyze_data():
 
    # 重新加载保存的汇总分析表文件
    summary_workbook = openpyxl.load_workbook('汇总分析表.xlsx')
    summary_sheet = summary_workbook.active


    default_fill = PatternFill(fill_type=None)
    default_font = Font()

    for row in summary_sheet.iter_rows():
        for cell in row:
            cell.fill = default_fill
            cell.font = default_font

    # 获取第一行的最大列数
    max_column = max(cell.column for cell in summary_sheet[1] if cell.value is not None)
    

    # 清空从max_column+1列到最后一列的内容
    for column in summary_sheet.iter_cols(min_col=max_column+1):
        for cell in column:
            cell.value = None


    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")


    max_count = 0
    other_values = []  # 初始化为空列表

    # 检查E列及其后续列的每个单元格是否为空
    for row_index in range(2, summary_sheet.max_row + 1):
        empty_cell_count = 0  # 变量
        for col_index in range(5, max_column + 1):
            # 检查当前单元格所在的行和列的第一个单元格是否同时不为空
            if (
                summary_sheet.cell(row=row_index, column=1).value
                and summary_sheet.cell(row=1, column=col_index).value
            ):
                # 检查当前单元格是否为空
                if summary_sheet.cell(row=row_index, column=col_index).value is None:
                    empty_cell_count += 1  # 计数数量
                    summary_sheet.cell(row=row_index, column=col_index).fill = red_fill

        if empty_cell_count>0:
            header_column = max_column + 1 # 向右遍历当前行的非空单元格
            next_cell = summary_sheet.cell(row=row_index, column=header_column)

            while next_cell.value is not None:
                header_column += 1
                next_cell = summary_sheet.cell(row=row_index, column=header_column)

            next_cell.value = "内容缺失，务必补齐源数据后重新汇总"

    summary_workbook.save('汇总分析表.xlsx')
              
    # 如果主观分包含符号，检查是否一致
    for row_index in range(2, summary_sheet.max_row + 1):
        has_number = False
        has_symbol = False

        for col_index in range(7, max_column + 1):
            cell_value = summary_sheet.cell(row=row_index, column=col_index).value

            if cell_value is not None:
                if isinstance(cell_value, (int, float)):
                    has_number = True #存在数字
                else:
                    has_symbol = True #存在符号

        if has_number and has_symbol:
            for col_index in range(7, max_column + 1):
                cell = summary_sheet.cell(row=row_index, column=col_index)
                cell.fill = red_fill  # 标记红色背景

        if has_number and has_symbol:
            header_column = max_column + 1 # 向右遍历当前行的非空单元格
            next_cell = summary_sheet.cell(row=row_index, column=header_column)

            while next_cell.value is not None:
                header_column += 1
                next_cell = summary_sheet.cell(row=row_index, column=header_column)

            next_cell.value = "同时存在数字和符号"
               
    summary_workbook.save('汇总分析表.xlsx')

    # 检查分值是否超限或小于0
    for row_index in range(2, summary_sheet.max_row + 1):
        empty_cell_count = 0  # 变量
        # 检查A单元格是否为空
        if summary_sheet.cell(row=row_index, column=1).value:
            
            # 检查F单元格到最大列数的当前行单元格
            
            for col_index in range(7, max_column + 1):
                cell_current = summary_sheet.cell(row=row_index, column=col_index)
                cell_e = summary_sheet.cell(row=row_index, column=5)
            
                # 检查单元格是否大于E或小于0
                if cell_current.value and re.match(r'^-?\d+(?:\.\d+)?$', str(cell_current.value)):
                    cell_value = float(cell_current.value)
                    if cell_value > cell_e.value or cell_value < 0:
                        empty_cell_count += 1  # 计数数量
                        cell_current.fill = red_fill  # 填充红色背景
                        cell_e.font = Font(color="FF0000", bold=True)  # 字体红色加粗

        if empty_cell_count>0:
            header_column = max_column + 1 # 向右遍历当前行的非空单元格
            next_cell = summary_sheet.cell(row=row_index, column=header_column)

            while next_cell.value is not None:
                header_column += 1
                next_cell = summary_sheet.cell(row=row_index, column=header_column)

            next_cell.value = "分值超限或小于0"
    summary_workbook.save('汇总分析表.xlsx')


    # 检查客观分是否一致
    for row_index in range(2, summary_sheet.max_row + 1):
        value_counts = {}
        row_cells = summary_sheet[row_index]

        if summary_sheet.cell(row=row_index, column=6).value == "客观分":
            for cell in row_cells[6:max_column]:
                value = cell.value
                if value:
                    if value not in value_counts:
                        value_counts[value] = 1
                    else:
                        value_counts[value] += 1

            if value_counts:
                max_count = max(value_counts.values())
                max_value = [value for value, count in value_counts.items() if count == max_count]
                other_values = [value for value, count in value_counts.items() if count != max_count]

            if max_count < (max_column - 6) :
                print(f"行号：{row_index}---其他客观值: {other_values}")
                for cell in row_cells[6:max_column]:
                    if cell.value is not None and cell.value in other_values:
                        cell.fill = red_fill # 填充红色背景
                        

            if value_counts and max_count < (max_column - 6) :
                header_column = max_column + 1 # 向右遍历当前行的非空单元格
                next_cell = summary_sheet.cell(row=row_index, column=header_column)

                while next_cell.value is not None:
                    header_column += 1
                    next_cell = summary_sheet.cell(row=row_index, column=header_column)

                next_cell.value = "客观分不一致"
    summary_workbook.save('汇总分析表.xlsx')





    # 检查主观分是否异常一致
    for row_index in range(2, summary_sheet.max_row + 1):
        value_counts = {}
        row_cells = summary_sheet[row_index]
        if summary_sheet.cell(row=row_index, column=6).value == "主观分":
            for cell in row_cells[6:max_column]:
                value = cell.value
                if value:
                    if isinstance(value, (int, float)):
                        if value not in value_counts:
                            value_counts[value] = 1
                        else:
                            value_counts[value] += 1

            if value_counts:
                max_count = max(value_counts.values())
                max_value = [value for value, count in value_counts.items() if count == max_count]
                other_values = [value for value, count in value_counts.items() if count != max_count]

            if max_count >= (max_column - 6)*0.6 :
                print(f"行号：{row_index}---最多数量主观分值: {max_value}---出现数量: {max_count}")
                for cell in row_cells[6:max_column]:
                    if cell.value is not None and cell.value in max_value:
                        cell.fill =PatternFill(start_color="FFCC99", fill_type="solid")
                        

        if value_counts and max_count >= (max_column - 6)*0.6 :
            header_column = max_column + 1 # 向右遍历当前行的非空单元格
            next_cell = summary_sheet.cell(row=row_index, column=header_column)

            while next_cell.value is not None:
                header_column += 1
                next_cell = summary_sheet.cell(row=row_index, column=header_column)

            next_cell.value = "主观分异常一致"
    summary_workbook.save('汇总分析表.xlsx')




    # 检查主观分偏离偏离30%以上
    for row_index in range(2, summary_sheet.max_row + 1):
        empty_cell_count = 0  # 变量
        if summary_sheet.cell(row=row_index, column=6).value == "主观分":
            values_to_compare = []
            for col_index in range(7, max_column + 1):
                cell_current = summary_sheet.cell(row=row_index, column=col_index)
                if cell_current.value and isinstance(cell_current.value, (int, float)):
                    values_to_compare.append(cell_current.value)

            if len(values_to_compare) > 0:
                average = sum(values_to_compare) / len(values_to_compare)
                for col_index in range(7, max_column + 1):
                    cell_current = summary_sheet.cell(row=row_index, column=col_index)
                    if cell_current.value and isinstance(cell_current.value, (int, float)):
                        if cell_current.value < average * 0.7 or cell_current.value > average * 1.3:
                            empty_cell_count += 1  # 计数数量
                            cell_current.fill = yellow_fill

        if empty_cell_count>0:
            header_column = max_column + 1 # 向右遍历当前行的非空单元格
            next_cell = summary_sheet.cell(row=row_index, column=header_column)

            while next_cell.value is not None:
                header_column += 1
                next_cell = summary_sheet.cell(row=row_index, column=header_column)

            next_cell.value = "主观分偏离平均值30%以上"
    summary_workbook.save('汇总分析表.xlsx')




    # 检查F列及其后续列的每个单元格是否等于0
    for row_index in range(2, summary_sheet.max_row + 1):
        empty_cell_count = 0  # 变量

        for col_index in range(7, max_column + 1):
            if summary_sheet.cell(row=row_index, column=col_index).value == 0:
                empty_cell_count += 1  # 计数数量
                cell = summary_sheet.cell(row=row_index, column=col_index)
                cell.font = Font(color="FF0000")  # 将字体颜色设置为红色

        if empty_cell_count>0:
            header_column = max_column + 1 # 向右遍历当前行的非空单元格
            next_cell = summary_sheet.cell(row=row_index, column=header_column)

            while next_cell.value is not None:
                header_column += 1
                next_cell = summary_sheet.cell(row=row_index, column=header_column)

            next_cell.value = "分值为0，请检查"



    # 保存修改后的汇总分析表
    summary_workbook.save('汇总分析表.xlsx')
    print("分值校验完成。")
    messagebox.showinfo("提示", "分值校验完成！")    



def title_button():

    # 重新加载保存的汇总分析表文件
    summary_workbook = openpyxl.load_workbook('汇总分析表.xlsx')
    summary_sheet = summary_workbook.active


    # 删除除第一个表以外的所有表
    for sheet in summary_workbook.sheetnames[1:]:
        summary_workbook.remove(summary_workbook[sheet])


    # 保存修改后的工作簿
    summary_workbook.save('汇总分析表.xlsx')


    # 重新加载保存的汇总分析表文件
    summary_workbook = openpyxl.load_workbook('汇总分析表.xlsx')
    summary_sheet = summary_workbook.active


    # 创建新的工作表
    second_sheet = summary_workbook.create_sheet('数据分析', 1)  # 在索引为1的位置创建新工作表


    # 初始化客观分总分和主观分总分变量
    objective_total_score = 0
    subjective_total_score = 0
    business_total_score = 0
    technology_service_total_score = 0

    # 获取第一个工作表的数据
    data_rows = summary_sheet.iter_rows(min_row=2, values_only=True)

    # 遍历数据行
    for row in data_rows:
        column_a_value = row[0]  # A列的值
        column_b_value = row[1]  # B列的值
        column_e_value = row[4]  # E列的值
        column_f_value = row[5]  # F列的值
        column_g_value = row[6]  # G列的值

        # 判断是否包含商务
        if column_a_value == summary_sheet['A2'].value and '商务' in column_b_value:
            business_total_score += column_e_value

        if column_a_value == summary_sheet['A2'].value and ('技术' in column_b_value or '服务' in column_b_value):
            technology_service_total_score += column_e_value

        # 判断是否满足客观分条件
        if column_a_value == summary_sheet['A2'].value and column_f_value == '客观分':
            objective_total_score += column_e_value

        # 判断是否满足主观分条件
        if column_a_value == summary_sheet['A2'].value and column_f_value == '主观分':
            subjective_total_score += column_e_value

    print("商务部分总分:", business_total_score)
    print("技术及服务部分总分:", technology_service_total_score)
    print("客观分总分:", objective_total_score)
    print("主观分总分:", subjective_total_score)

    total_score = objective_total_score + subjective_total_score

    # 设置第二个工作表的D、E列表头为主观分总分分值、客观分总分分值
    second_sheet.cell(row=1, column=1).value = f"商务、技术及服务总得分\n（{total_score}分）"
    second_sheet.cell(row=1, column=2).value = f"技术及服务部分\n（{technology_service_total_score}分）"
    second_sheet.cell(row=1, column=3).value = f"商务部分\n（{business_total_score}分）"
    second_sheet.cell(row=1, column=4).value = "公司名称"
    second_sheet.cell(row=1, column=5).value = f"客观分\n（{objective_total_score}分）"
    second_sheet.cell(row=1, column=6).value = f"主观分\n（{subjective_total_score}分）"
    second_sheet.cell(row=1, column=7).value = "客观分-失分项"
    second_sheet.cell(row=1, column=8).value = "主观分-失分项"



    # 获取第一行的最大列数
    max_column = 0
    for cell in summary_sheet[1]:
        if cell.value is not None:
            max_column = cell.column


    # 创建字典来存储每个公司的得分和失分项
    company_scores = {}

    # 从第二行开始遍历每一行
    for row_index, row in enumerate(summary_sheet.iter_rows(min_row=2, values_only=True), start=2):
        company_name = row[0]  # 公司名称在A列
        numeric_scores = [value for value in row[6:max_column] if isinstance(value, (int, float))]
        average_score = sum(numeric_scores) / len(numeric_scores) if numeric_scores else 0
        lost_factors1 = []  # 保存客观失分项
        lost_factors2 = []  # 保存主观失分项

        # 检查公司名称是否已存在于字典中，如果不存在则添加新的键值对
        if company_name not in company_scores:
            company_scores[company_name] = {
                '商务、技术及服务总得分': 0,
                '客观分得分': 0,
                '主观分得分': 0,
                '商务部分得分': 0,
                '技术及服务部分得分': 0,
                '客观分失分项': lost_factors1,  # 保存客观分失分项
                '主观分失分项': lost_factors2,  # 保存主观分失分项
                '客观分评审因素得分': {},  # 客观分评审因素得分字典
                '主观分评审因素得分': {},  # 主观分评审因素得分字典
                '客观分评审因素排名': {},  # 客观分评审因素排名字典
                '主观分评审因素排名': {}  # 主观分评审因素排名字典
            }

        # 更新每个公司的得分和失分项
        company_scores[company_name]['商务、技术及服务总得分'] += average_score

        if row[5] == '主观分':
            company_scores[company_name]['主观分得分'] += average_score

        if row[5] == '客观分':
            company_scores[company_name]['客观分得分'] += average_score

        if '商务' in row[1]:
            company_scores[company_name]['商务部分得分'] += average_score

        if '技术' in row[1] or '服务' in row[1]:
            company_scores[company_name]['技术及服务部分得分'] += average_score

        # 当平均分小于E列分值的60%时，将失分项添加到对应公司的列表中
        if row[5] == '客观分' and average_score < row[4] * 0.6:
            company_scores[company_name]['客观分失分项'].append(row[2])

        if row[5] == '主观分' and average_score < row[4] * 0.6:
            company_scores[company_name]['主观分失分项'].append(row[2])

        # 将评审因素得分添加到对应公司的字典中
        if row[5] == '客观分':
            company_scores[company_name]['客观分评审因素得分'][row[2]] = average_score

        if row[5] == '主观分':
            company_scores[company_name]['主观分评审因素得分'][row[2]] = average_score

    # 对每个公司的客观分评审因素得分进行排名
    for company_name, scores in company_scores.items():
        objective_scores = scores['客观分评审因素得分']
        sorted_objective_scores = sorted(objective_scores.items(), key=lambda x: x[1], reverse=True)
        objective_ranks = {factor: rank + 1 for rank, (factor, _) in enumerate(sorted_objective_scores)}
        company_scores[company_name]['客观分评审因素排名'] = objective_ranks

    # 对每个公司的主观分评审因素得分进行排名
    for company_name, scores in company_scores.items():
        subjective_scores = scores['主观分评审因素得分']
        sorted_subjective_scores = sorted(subjective_scores.items(), key=lambda x: x[1], reverse=True)
        subjective_ranks = {factor: rank + 1 for rank, (factor, _) in enumerate(sorted_subjective_scores)}
        company_scores[company_name]['主观分评审因素排名'] = subjective_ranks

    # 将公司得分和评审因素排名写入表2的相应列
    for row_index, (company_name, scores) in enumerate(company_scores.items(), start=2):
        second_sheet.cell(row=row_index, column=1).value = scores['商务、技术及服务总得分']
        second_sheet.cell(row=row_index, column=2).value = scores['技术及服务部分得分']
        second_sheet.cell(row=row_index, column=3).value = scores['商务部分得分']
        second_sheet.cell(row=row_index, column=4).value = company_name
        second_sheet.cell(row=row_index, column=5).value = scores['客观分得分']
        second_sheet.cell(row=row_index, column=6).value = scores['主观分得分']


        # 获取客观失分因素列表
        lost_factors1 = scores['客观分失分项']
        if lost_factors1:
            lost_factors_str = '、'.join(lost_factors1)
            # 获取客观分评审因素排名在80%以外的评审因素列表
            objective_factors = [factor for factor, rank in scores['客观分评审因素排名'].items() if rank > math.ceil(len(scores['客观分评审因素排名']) * 0.2)]
            objective_factors_str = '、'.join(objective_factors)

            combined_str = f"【{lost_factors_str}】等评审因素失分较多；\n【{objective_factors_str}】等评审因素得分排名较低。"
            second_sheet.cell(row=row_index, column=7).value = combined_str  # 将两个内容合并后写入G列
      
        # 获取主观分失分因素列表
        lost_factors2 = scores['主观分失分项']
        if lost_factors2:
            lost_factors_str = '、'.join(lost_factors2)
            # 获取主观分评审因素排名在80%以外的评审因素列表
            subjective_factors = [factor for factor, rank in scores['主观分评审因素排名'].items() if rank > math.ceil(len(scores['主观分评审因素排名']) * 0.2)]
            subjective_factors_str = '、'.join(subjective_factors)

            combined_str = f"【{lost_factors_str}】等评审因素失分较多；\n【{subjective_factors_str}】等评审因素得分排名较低。"
            second_sheet.cell(row=row_index, column=8).value = combined_str  # 将两个内容合并后写入H列

    # 获取最大行数和列数
    max_row = second_sheet.max_row
    max_column = second_sheet.max_column

    # 设置所有单元格的对齐方式和换行\字体为宋体，字号为10
    for row in second_sheet.iter_rows():
        for cell in row:
            cell.font = Font(name='宋体', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


    # 设置 列的宽度
    second_sheet.column_dimensions['A'].width = 8
    second_sheet.column_dimensions['B'].width = 15
    second_sheet.column_dimensions['C'].width = 8
    second_sheet.column_dimensions['D'].width = 18
    second_sheet.column_dimensions['E'].width = 8
    second_sheet.column_dimensions['F'].width = 8
    second_sheet.column_dimensions['G'].width = 50
    second_sheet.column_dimensions['H'].width = 50

    # 创建边框样式
    border_style = Side(border_style="thin", color="000000")  # 选取细线样式和黑色颜色
    cell_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)



    # 遍历 A1 到最大列和最大行的单元格，并设置边框
    for row in second_sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)


    # 保存修改后的工作簿
    summary_workbook.save('汇总分析表.xlsx')
    messagebox.showinfo("提示", "数据分析完成！") 


# 横向反向更新
def horizontal_update():
    # 打开汇总分析表.xlsx和工作簿1.xlsx
    wb_summary = openpyxl.load_workbook('汇总分析表.xlsx', data_only=True)
    wb_workbook1 = openpyxl.load_workbook('工作簿1.xlsx')

    # 获取汇总分析表的活动工作表
    summary_sheet = wb_summary.active

    # 获取第一行的最大列
    max_column = max(cell.column for cell in summary_sheet[1] if cell.value is not None)
    start_column = column_index_from_string('G')  # 'G'列对应的列数

    # 获取所有需要处理的列字母，从'G'列到最大列
    column_letters = [get_column_letter(i) for i in range(start_column, max_column + 1)]

    # 遍历每一列
    for letter in column_letters:
        # 获取汇总分析表.xlsx中的sheet工作表的当前列的第一个单元格的内容
        column1_value = summary_sheet[f'{letter}1'].value

        # 获取工作簿1.xlsx中的当前工作表
        workbook1_sheet = wb_workbook1[column1_value]

        # 遍历汇总分析表.xlsx的sheet工作表的当前列，从第2个单元格开始
        for row in summary_sheet.iter_rows(min_row=2, min_col=column_index_from_string(letter), max_col=column_index_from_string(letter)):
            column_value = row[0].value  # 获取实际存储的值
            number_format = summary_sheet[f'{letter}{row[0].row}'].number_format

            if column_value is not None:
                column_value = Decimal(column_value).quantize(Decimal('0.0000000000'), rounding=ROUND_HALF_UP)
            else:
                continue
            
            # 获取汇总分析表.xlsx中的sheet工作表对应的A列和C列的内容
            a_value = summary_sheet.cell(row=row[0].row, column=1).value
            c_value = summary_sheet.cell(row=row[0].row, column=3).value

            # 在工作簿1.xlsx的当前工作表中查找匹配的公司名称和评审项
            company_row = None
            company_column = None
            review_row = None
            review_column = None
            for cell in workbook1_sheet.iter_rows():
                for cell in cell:
                    if cell.value == a_value:
                        company_row = cell.row
                        company_column = cell.column
                    elif cell.value == c_value:
                        review_row = cell.row
                        review_column = cell.column

            if company_row is None or review_row is None:
                continue

            # 获取目标单元格的原有值
            target_cell = workbook1_sheet.cell(company_row, column=review_column)
            old_value = target_cell.value

            # 如果新的值和旧的值不同，就执行更新操作
            if column_value != old_value:
                target_cell.value = float(column_value)  # 更新时转为浮点数
                target_cell.number_format = number_format  # 确保精度一致
                print(f'【{column1_value}】表【{get_column_letter(company_column)}{review_row}】单元格 值 {old_value} 更新为 {column_value}')

        # 保存工作簿1.xlsx
        wb_workbook1.save('工作簿1.xlsx')
    messagebox.showinfo("提示", "横版反向更新完成！")

# 纵向反向更新
def vertical_update():
    # 打开汇总分析表.xlsx和工作簿1.xlsx
    wb_summary = openpyxl.load_workbook('汇总分析表.xlsx', data_only=True)
    wb_workbook1 = openpyxl.load_workbook('工作簿1.xlsx')

    # 获取汇总分析表的活动工作表
    summary_sheet = wb_summary.active

    # 获取第一行的最大列
    max_column = max(cell.column for cell in summary_sheet[1] if cell.value is not None)
    start_column = column_index_from_string('G')  # 'G'列对应的列数

    # 获取所有需要处理的列字母，从'G'列到最大列
    column_letters = [get_column_letter(i) for i in range(start_column, max_column + 1)]

    # 遍历每一列
    for letter in column_letters:
        # 获取汇总分析表.xlsx中的sheet工作表的当前列的第一个单元格的内容
        column1_value = summary_sheet[f'{letter}1'].value

        # 获取工作簿1.xlsx中的当前工作表
        workbook1_sheet = wb_workbook1[column1_value]

        # 遍历汇总分析表.xlsx的sheet工作表的当前列，从第2个单元格开始
        for row in summary_sheet.iter_rows(min_row=2, min_col=column_index_from_string(letter), max_col=column_index_from_string(letter)):
            column_value = row[0].value  # 获取实际存储的值
            number_format = summary_sheet[f'{letter}{row[0].row}'].number_format

            if column_value is not None:
                column_value = Decimal(column_value).quantize(Decimal('0.0000000000'), rounding=ROUND_HALF_UP)
            else:
                continue
            
            # 获取汇总分析表.xlsx中的sheet工作表对应的A列和C列的内容
            a_value = summary_sheet.cell(row=row[0].row, column=1).value
            c_value = summary_sheet.cell(row=row[0].row, column=3).value

            # 在工作簿1.xlsx的当前工作表中查找匹配的公司名称和评审项
            company_row = None
            company_column = None
            review_row = None
            review_column = None
            for cell in workbook1_sheet.iter_rows():
                for cell in cell:
                    if cell.value == a_value:
                        company_row = cell.row
                        company_column = cell.column
                    elif cell.value == c_value:
                        review_row = cell.row
                        review_column = cell.column

            if company_row is None or review_row is None:
                continue

            # 获取目标单元格的原有值
            target_cell = workbook1_sheet.cell(row=review_row, column=company_column)
            old_value = target_cell.value

            # 如果新的值和旧的值不同，就执行更新操作
            if column_value != old_value:
                target_cell.value = float(column_value)  # 更新时转为浮点数
                target_cell.number_format = number_format  # 确保精度一致
                print(f'【{column1_value}】表【{get_column_letter(company_column)}{review_row}】单元格 值 {old_value} 更新为 {column_value}')

        # 保存工作簿1.xlsx
        wb_workbook1.save('工作簿1.xlsx')
    messagebox.showinfo("提示", "纵版反向更新完成！")
    

# 创建主窗口
def price_calculation_main():
    import matplotlib.pyplot as plt
    from decimal import getcontext
    from itertools import cycle

    plt.rcParams['font.family'] = ['SimSun', 'Microsoft YaHei', 'Arial Unicode MS', 'sans-serif']

    getcontext().prec = 28
    getcontext().rounding = ROUND_HALF_UP

    def round_half_up(n, decimals=2):
        return float(Decimal(n).quantize(Decimal('1.' + '0' * decimals), rounding=ROUND_HALF_UP))

    def calculate_result():
        entry_result.delete("1.0", tk.END)
        result_text.delete("1.0", tk.END)

        input_items = [
            (entry_a.get(), "价格满分"),
            (entry_q.get(), "得分比率Q"),
            (entry_high_lambda.get(), "高于基准价时λ取值"),
            (entry_low_lambda.get(), "低于基准价时λ取值"),
            (entry_positive_b.get(), "正偏离X以上异常"),
            (entry_negative_c.get(), "负偏离Y以下异常"),
        ]
        for value, label in input_items:
            if not value.strip():
                entry_result.insert(tk.END, f"{label}未填写，无法计算。\n")
                result_text.insert(tk.END, f"{label}未填写，无法计算。\n")
                return

        def parse_decimal_value(raw_value):
            value = raw_value.strip()
            if value.endswith('%'):
                return Decimal(value[:-1]) / 100
            return Decimal(value)

        try:
            a = Decimal(entry_a.get().strip())
            q = Decimal(entry_q.get().strip())
            high_lambda = Decimal(entry_high_lambda.get().strip())
            low_lambda = Decimal(entry_low_lambda.get().strip())
            positive_b = parse_decimal_value(entry_positive_b.get())
            negative_c = parse_decimal_value(entry_negative_c.get())
        except Exception:
            entry_result.insert(tk.END, "输入格式有误，无法计算。\n")
            result_text.insert(tk.END, "输入格式有误，无法计算。\n")
            return

        prices_text = entry_price.get("1.0", tk.END)
        prices = []

        for line in prices_text.split("\n"):
            line = line.strip()
            if line:
                match = re.search(r"(\d+(\.\d+)?)(%?)", line)
                if match:
                    value = Decimal(match.group(1))
                    if match.group(3) == '%':
                        value /= 100
                    prices.append(value)
                else:
                    entry_result.insert(tk.END, f"价格输入格式错误：{line}\n")
                    result_text.insert(tk.END, f"价格输入格式错误：{line}\n")
                    return

        if not prices:
            entry_result.insert(tk.END, "未输入有效价格，无法计算。\n")
            result_text.insert(tk.END, "未输入有效价格，无法计算。\n")
            return

        average_price = sum(prices, Decimal("0")) / len(prices)
        entry_result.insert(tk.END, f"所有价格的平均价：{average_price:.8}\n\n")
        result_text.insert(tk.END, f"所有价格的平均价: {average_price:.8}\n\n")

        exceptional_prices = []
        for price in prices:
            if price > average_price * (1 + positive_b) or price < average_price * (1 - negative_c):
                exceptional_prices.append(price)
        entry_result.insert(tk.END, "异常报价判定:\n")
        result_text.insert(tk.END, "异常报价判定:\n")
        for price in prices:
            deviation = (price - average_price) / average_price
            if price in exceptional_prices:
                entry_result.insert(tk.END, f"报价{price:>12} |偏离{deviation:>8.2%}→异常\n")
                result_text.insert(tk.END, f"报价{price:>12} |偏离{deviation:>8.2%}→异常\n")
            else:
                entry_result.insert(tk.END, f"报价{price:>12} |偏离{deviation:>8.2%}\n")
                result_text.insert(tk.END, f"报价{price:>12} |偏离{deviation:>8.2%}\n")
        entry_result.insert(tk.END, "\n")
        result_text.insert(tk.END, "\n")

        non_exceptional_prices_for_baseline_method = [price for price in prices if price not in exceptional_prices]
        non_exceptional_prices_for_interpolation_method = [price for price in prices if price not in exceptional_prices]

        if len(non_exceptional_prices_for_baseline_method) < 1:
            baseline_price = average_price
        else:
            baseline_price = sum(non_exceptional_prices_for_baseline_method) / len(non_exceptional_prices_for_baseline_method)

        if len(non_exceptional_prices_for_interpolation_method) < 2:
            valid_max_price = max(prices)
            valid_min_price = min(prices)
        else:
            valid_max_price = max(non_exceptional_prices_for_interpolation_method)
            valid_min_price = min(non_exceptional_prices_for_interpolation_method)

        entry_result.insert(tk.END, f"基准价：{baseline_price:.8}\n\n\n")
        result_text.insert(tk.END, f"有效最高报价: {valid_max_price}\n")
        result_text.insert(tk.END, f"有效最低报价: {valid_min_price}\n\n")

        entry_result.insert(tk.END, "每个报价及对应的价格得分：\n")
        price_scores = []
        for price in prices:
            score = a * q * (1 - (price - baseline_price) / baseline_price * (high_lambda if price > baseline_price else low_lambda))
            score = round_half_up(score)
            score = max(min(a, score), 0)
            price_scores.append(score)
        for price, score in zip(prices, price_scores):
            entry_result.insert(tk.END, f"报价{price:>12} |得分{score:>6.2f}\n")

        result_text.insert(tk.END, "每个报价对应的价格得分:\n")
        price_scores1 = []
        for price in prices:
            score1 = (valid_max_price + valid_min_price - price) / valid_max_price * a
            score1 = round_half_up(score1)
            score1 = max(min(a, score1), 0)
            price_scores1.append(score1)
        for price, score1 in zip(prices, price_scores1):
            result_text.insert(tk.END, f"报价{price:>12} |得分{score1:>6.2f}\n")

    color_options = ['blue', 'red', 'green', 'purple', 'orange', 'brown', 'black', 'gray', 'yellow', 'cyan', 'magenta', 'Lavender']
    color_cycle = cycle(color_options)

    def draw_scatter_plot():
        prices_text = entry_price.get("1.0", tk.END)
        raw_prices = []

        for line in prices_text.split("\n"):
            line = line.strip()
            if line:
                match = re.search(r"(\d+(\.\d+)?)(%?)", line)
                if match:
                    value = float(match.group(1))
                    if match.group(3) == '%':
                        value /= 100
                    raw_prices.append(value)

        price_scores_text = entry_result.get("1.0", tk.END)
        price_scores = [float(score.split("|得分")[1]) for score in price_scores_text.split("\n") if score.strip() and "|得分" in score]

        price_scores1_text = result_text.get("1.0", tk.END)
        price_scores1 = [float(score.split("|得分")[1]) for score in price_scores1_text.split("\n") if score.strip() and "|得分" in score]

        if not raw_prices:
            messagebox.showwarning("提示", "未检测到有效价格，无法绘图。")
            return

        if len(price_scores) != len(raw_prices) or len(price_scores1) != len(raw_prices):
            messagebox.showwarning("提示", "价格数量与得分数量不一致，请先点击“开始计算”后再绘图。")
            return

        plt.clf()

        if raw_prices and price_scores:
            baseline_pairs = sorted(zip(raw_prices, price_scores))
            baseline_prices, baseline_scores = zip(*baseline_pairs)
            color = next(color_cycle)
            plt.plot(baseline_prices, baseline_scores, color=color, alpha=0.5, label='HB经修正的基准价法')
            plt.scatter(baseline_prices, baseline_scores, color=color, alpha=0.5)

        if raw_prices and price_scores1:
            interpolation_pairs = sorted(zip(raw_prices, price_scores1))
            interpolation_prices, interpolation_scores = zip(*interpolation_pairs)
            color = next(color_cycle)
            plt.plot(interpolation_prices, interpolation_scores, color=color, alpha=0.5, label='经修正的直线内插法')
            plt.scatter(interpolation_prices, interpolation_scores, color=color, alpha=0.5)

        if raw_prices and (price_scores or price_scores1):
            plt.xlabel('价格')
            plt.ylabel('得分')
            plt.title('价格得分散点图')
            plt.legend()
            plt.show()

    window = tk.Tk()
    window.title("HB价格测算")
    window.geometry("940x320")

    parameters_frame = tk.Frame(window)
    parameters_frame.grid(row=0, column=0, padx=10, sticky="n")

    label_a = tk.Label(parameters_frame, text="价格满分")
    label_a.grid(row=0, column=0, sticky="w")
    entry_a = tk.Entry(parameters_frame)
    entry_a.grid(row=1, column=0)

    label_q = tk.Label(parameters_frame, text="得分比率Q")
    label_q.grid(row=2, column=0, sticky="w")
    entry_q = tk.Entry(parameters_frame)
    entry_q.insert(tk.END, "0.9")
    entry_q.grid(row=3, column=0)

    label_high_lambda = tk.Label(parameters_frame, text="高于基准价时λ取值")
    label_high_lambda.grid(row=4, column=0, sticky="w")
    entry_high_lambda = tk.Entry(parameters_frame)
    entry_high_lambda.insert(tk.END, "1")
    entry_high_lambda.grid(row=5, column=0)

    label_low_lambda = tk.Label(parameters_frame, text="低于基准价时λ取值")
    label_low_lambda.grid(row=6, column=0, sticky="w")
    entry_low_lambda = tk.Entry(parameters_frame)
    entry_low_lambda.insert(tk.END, "0.5")
    entry_low_lambda.grid(row=7, column=0)

    label_positive_b = tk.Label(parameters_frame, text="正偏离X以上异常")
    label_positive_b.grid(row=8, column=0, sticky="w")
    entry_positive_b = tk.Entry(parameters_frame)
    entry_positive_b.grid(row=9, column=0)

    label_negative_c = tk.Label(parameters_frame, text="负偏离Y以下异常")
    label_negative_c.grid(row=10, column=0, sticky="w")
    entry_negative_c = tk.Entry(parameters_frame)
    entry_negative_c.grid(row=11, column=0)

    button_calculate = tk.Button(parameters_frame, text="开始计算", command=calculate_result)
    button_calculate.grid(row=12, column=0, pady=10, sticky="w")

    price_frame = tk.Frame(window)
    price_frame.grid(row=0, column=1, padx=10, sticky="n")

    label_price = tk.Label(price_frame, text="逐行输入价格")
    label_price.grid(row=0, column=0, sticky="w")
    entry_price = tk.Text(price_frame, height=21, width=20)
    entry_price.grid(row=1, column=0, sticky="w")

    result_frame = tk.Frame(window)
    result_frame.grid(row=0, column=2, padx=10, sticky="n")
    label_result = tk.Label(result_frame, text="经修正的基准价结果输出")
    label_result.grid(row=0, column=0, sticky="w")
    entry_result = tk.Text(result_frame, height=21, width=40)
    entry_result.grid(row=1, column=0, sticky="w")
    entry_result.insert(tk.END, "【经修正的基准价法】\n\n1.价格得分=价格满分*得分比率Q*(1-(投标报价-基准价)/ 基准价*λ)；\n\n2.投标报价高于或低于基准价时λ分别取值；\n\n3.异常报价：所有投标人报价的算术平均值正偏离X以上，负偏离Y以下的投标报价为异常报价；\n\n4.基准价：所有非异常投标报价的算术平均值；\n\n5.若投标人价格均为异常，则取所有投标价格的算术平均价作为基准价；\n\n6.所有投标人价格得分（包含异常价格）均采用此公式计算，最高加至满分，最低扣至0分。")

    result_frame_interpolation = tk.Frame(window)
    result_frame_interpolation.grid(row=0, column=3, padx=10, sticky="n")
    label_result_interpolation = tk.Label(result_frame_interpolation, text="经修正的直线内插法结果输出")
    label_result_interpolation.grid(row=0, column=0, sticky="w")
    result_text = tk.Text(result_frame_interpolation, height=21, width=40)
    result_text.grid(row=1, column=0, sticky="w")
    result_text.insert(tk.END, "【经修正的直线内插法】\n\n1.价格得分=（（有效最高报价+有效最低报价）-投标报价)/有效最高报价*价格满分；\n\n2.异常报价：所有投标人报价的算术平均值正偏离X以上、负偏离Y以下的投标报价为异常报价；\n\n3.有效最高报价和有效最低报价为剔除异常报价后的所有报价中的最高、最低报价；\n\n4.若剔除异常报价后剩余报价少于2个，则不再作异常报价认定，直接以所有投标人报价中的最高报价和最低报价代入计算公式计算；\n\n5.异常报价投标人得分同样按以上计算公式计算；\n\n6.所有投标人价格得分（包含异常价格）均采用此公式计算，最高得满分，最低得0分。")

    result_frame_plot = tk.Frame(window)
    result_frame_plot.grid(row=0, column=4, padx=10, sticky="n")
    label_result_plot = tk.Label(result_frame_plot, text="散点图")
    label_result_plot.grid(row=0, column=0, sticky="w")

    button_plot = tk.Button(parameters_frame, text="绘制散点图", command=draw_scatter_plot)
    button_plot.grid(row=12, column=0, pady=10, sticky="e")

    window.mainloop()


def open_price_calculation():
    try:
        window.destroy()
        price_calculation_main()
    except Exception as e:
        messagebox.showerror("提示", f"打开价格测算失败：{e}")


window = tk.Tk()
window.title("汇总分析")
window.geometry("255x165")

# 创建按钮容器框架
button_frame = tk.Frame(window)
button_frame.pack(pady=10)

# 创建横版数据汇总按钮
summarize_horizontal_button = tk.Button(button_frame, text="横版数据汇总", command=summarize_data_horizontal, width=15, height=1)
summarize_horizontal_button.grid(row=0, column=0, padx=5, pady=3)

# 创建纵版数据汇总按钮
summarize_vertical_button = tk.Button(button_frame, text="纵版数据汇总", command=summarize_data_vertical, width=15, height=1)
summarize_vertical_button.grid(row=1, column=0, padx=5, pady=3)

# 创建主客观分析按钮
subjective_analysis_button = tk.Button(button_frame, text="主客观分判断", command=subjective_analysis, width=15, height=1)
subjective_analysis_button.grid(row=0, column=1, padx=5, pady=3)

# 创建分值校验按钮
analyze_button = tk.Button(button_frame, text="分值校验", command=analyze_data, width=15, height=1)
analyze_button.grid(row=1, column=1, padx=5, pady=3)

# 创建数据分析按钮
title_button = tk.Button(button_frame, text="数据分析", command=title_button, width=15, height=1)
title_button.grid(row=2, column=1, padx=5, pady=3)

# 创建横版反向更新按钮
horizontal_update = tk.Button(button_frame, text="横版反向更新", command=horizontal_update, width=15, height=1)
horizontal_update.grid(row=2, column=0, padx=5, pady=3)

# 创建纵版反向更新按钮
vertical_update = tk.Button(button_frame, text="纵版反向更新", command=vertical_update, width=15, height=1)
vertical_update.grid(row=3, column=0, padx=5, pady=3)

price_calc_button = tk.Button(button_frame, text="价格测算", command=open_price_calculation, width=15, height=1)
price_calc_button.grid(row=3, column=1, padx=5, pady=3)


window.mainloop()
